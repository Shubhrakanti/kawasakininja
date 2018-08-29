import pandas as pd
from openpyxl import load_workbook
from xlrd import open_workbook
from xlutils.copy import copy
from math import sin, cos, sqrt, atan2, radians

#read excel sheets with all the data
AMC_all_df = pd.read_excel("Geocoded Data/AMC Geocoded Data.xlsx")
CNK_all_df = pd.read_excel("Geocoded Data/CNK Geocoded Data.xlsx")
RGC_all_df = pd.read_excel("Geocoded Data/RGC Geocoded Data.xlsx")

#seperate the premium loctations for each brand
AMC_premium_df = AMC_all_df[AMC_all_df["Seating"] == 1.0]
CNK_premium_df = CNK_all_df[CNK_all_df["Loungers"] == 1.0]
RGC_premium_df = RGC_all_df[RGC_all_df["Loungers"] == 1.0]

#seperate the non premium locations for each brand
AMC_reg_df = AMC_all_df[AMC_all_df["Seating"] != 1.0]
CNK_reg_df = CNK_all_df[CNK_all_df["Loungers"] != 1.0]
RGC_reg_df = RGC_all_df[RGC_all_df["Loungers"] != 1.0]


#formula to get dinstance beetween two geo coordiantes
def get_distance(lat1,lon1,lat2,lon2):
    R = 3959

    lat1 = radians(lat1)
    lon1 = radians(lon1)
    lat2 = radians(lat2)
    lon2 = radians(lon2)

    dlon = lon2 - lon1
    dlat = lat2 - lat1

    a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
    c = 2 * atan2(sqrt(a), sqrt(1 - a))

    distance = R * c
    return distance


#Functions calculate the percent of premium locations that have at least "num_loc"
#number of premium location from any company in a "rad" mile radius
def AMC_percent_in_rad_premium(num_loc,rad):

    num_locations_with_intersects = []
    for amc_location in AMC_premium_df.iterrows():
        current_intersects = 0
        check_cnk = True
        for rgc_location in RGC_premium_df.iterrows():
            if get_distance(rgc_location[1]["lats"],rgc_location[1]["longs"],amc_location[1]["lats"],amc_location[1]["longs"]) <= rad:
                current_intersects += 1
            if current_intersects == num_loc:
                num_locations_with_intersects.append(amc_location[1]["Address"])
                check_cnk = False
                break
        if check_cnk:
            for cnk_location in CNK_premium_df.iterrows():
                if get_distance(cnk_location[1]["lats"],cnk_location[1]["longs"],amc_location[1]["lats"],amc_location[1]["longs"]) <= rad:
                    current_intersects += 1
                if current_intersects == num_loc:
                    num_locations_with_intersects.append(amc_location[1]["Address"])
                    break
    return len(num_locations_with_intersects)/AMC_premium_df["Address"].count(), len(num_locations_with_intersects)

def AMC_percent_in_rad_premium_RGC(num_loc,rad, premium):

    num_locations_with_intersects = []
    for amc_location in AMC_premium_df.iterrows():
        current_intersects = 0
        #check_cnk = True
        lst = RGC_premium_df.iterrows()
        if not premium:
            lst = RGC_reg_df.iterrows()
        for rgc_location in lst:
            if get_distance(rgc_location[1]["lats"],rgc_location[1]["longs"],amc_location[1]["lats"],amc_location[1]["longs"]) <= rad:
                current_intersects += 1
            if current_intersects == num_loc:
                num_locations_with_intersects.append(amc_location[1]["Address"])
                #check_cnk = False
                break
    return len(num_locations_with_intersects)/AMC_premium_df["Address"].count(), len(num_locations_with_intersects)

def AMC_percent_in_rad_premium_CNK(num_loc,rad, premium):

    num_locations_with_intersects = []
    for amc_location in AMC_premium_df.iterrows():
        current_intersects = 0
        lst = CNK_premium_df.iterrows()
        if not premium:
            lst = CNK_reg_df.iterrows()
        for cnk_location in lst:
            if get_distance(cnk_location[1]["lats"],cnk_location[1]["longs"],amc_location[1]["lats"],amc_location[1]["longs"]) <= rad:
                current_intersects += 1
            if current_intersects == num_loc:
                num_locations_with_intersects.append(amc_location[1]["Address"])
                break
    return len(num_locations_with_intersects)/AMC_premium_df["Address"].count(), len(num_locations_with_intersects)


def AMC_percent_in_rad_premium_Both(num_loc,rad, premium):

    num_locations_with_intersects = []
    for amc_location in AMC_premium_df.iterrows():
        current_intersects = 0
        #check_cnk = True
        lst1 = RGC_premium_df.iterrows()
        lst2 = CNK_premium_df.iterrows()
        if not premium:
            lst1 = RGC_reg_df.iterrows()
            lst2 = CNK_reg_df.iterrows()

        for rgc_location in lst1:
            if get_distance(rgc_location[1]["lats"],rgc_location[1]["longs"],amc_location[1]["lats"],amc_location[1]["longs"]) <= rad:
                for cnk_location in lst2:
                    if get_distance(cnk_location[1]["lats"],cnk_location[1]["longs"],amc_location[1]["lats"],amc_location[1]["longs"]) <= rad:
                        current_intersects += 1
            if current_intersects == num_loc:
                num_locations_with_intersects.append(amc_location[1]["Address"])
                #check_cnk = False
                break
    return len(num_locations_with_intersects)/AMC_premium_df["Address"].count(), len(num_locations_with_intersects)

def AMC_percent_in_rad_premium_None(num_loc,rad, premium):

    num_locations_with_intersects = []
    for amc_location in AMC_premium_df.iterrows():
        current_intersects = 0
        #check_cnk = True
        lst1 = RGC_premium_df.iterrows()
        lst2 = CNK_premium_df.iterrows()
        if not premium:
            lst1 = RGC_reg_df.iterrows()
            lst2 = CNK_reg_df.iterrows()
        for rgc_location in lst1:
            if not get_distance(rgc_location[1]["lats"],rgc_location[1]["longs"],amc_location[1]["lats"],amc_location[1]["longs"]) <= rad:
                for cnk_location in lst2:
                    if not get_distance(cnk_location[1]["lats"],cnk_location[1]["longs"],amc_location[1]["lats"],amc_location[1]["longs"]) <= rad:
                        current_intersects += 1
            if current_intersects == num_loc:
                num_locations_with_intersects.append(amc_location[1]["Address"])
                #check_cnk = False
                break
    return len(num_locations_with_intersects)/AMC_premium_df["Address"].count(), len(num_locations_with_intersects)

def RGC_percent_in_rad_premium(num_loc,rad):

    num_locations_with_intersects = []
    for rgc_location in RGC_premium_df.iterrows():
        current_intersects = 0
        check_cnk = True
        for amc_location in AMC_premium_df.iterrows():
            if get_distance(rgc_location[1]["lats"],rgc_location[1]["longs"],amc_location[1]["lats"],amc_location[1]["longs"]) <= rad:
                current_intersects += 1
            if current_intersects == num_loc:
                num_locations_with_intersects.append(rgc_location[1]["Address"])
                check_cnk = False
                break
        if check_cnk:
            for cnk_location in CNK_premium_df.iterrows():
                if get_distance(cnk_location[1]["lats"],cnk_location[1]["longs"],rgc_location[1]["lats"],rgc_location[1]["longs"]) <= rad:
                    current_intersects += 1
                if current_intersects == num_loc:
                    num_locations_with_intersects.append(rgc_location[1]["Address"])
                    break
    return len(num_locations_with_intersects)/RGC_premium_df["Address"].count(), len(num_locations_with_intersects)

def RGC_percent_in_rad_premium_AMC(num_loc,rad, premium):

    num_locations_with_intersects = []
    for rgc_location in RGC_premium_df.iterrows():
        current_intersects = 0
        lst = AMC_premium_df.iterrows()
        if not premium:
            lst = AMC_reg_df.iterrows()
        for amc_location in lst:
            if get_distance(rgc_location[1]["lats"],rgc_location[1]["longs"],amc_location[1]["lats"],amc_location[1]["longs"]) <= rad:
                current_intersects += 1
            if current_intersects == num_loc:
                num_locations_with_intersects.append(rgc_location[1]["Address"])
                break
    return len(num_locations_with_intersects)/RGC_premium_df["Address"].count(), len(num_locations_with_intersects)

def RGC_percent_in_rad_premium_CNK(num_loc,rad, premium):

    num_locations_with_intersects = []
    for rgc_location in RGC_premium_df.iterrows():
        current_intersects = 0
        lst = CNK_premium_df.iterrows()
        if not premium:
            lst = CNK_reg_df.iterrows()
        for cnk_location in lst:
            if get_distance(cnk_location[1]["lats"],cnk_location[1]["longs"],rgc_location[1]["lats"],rgc_location[1]["longs"]) <= rad:
                current_intersects += 1
            if current_intersects == num_loc:
                num_locations_with_intersects.append(rgc_location[1]["Address"])
                break
    return len(num_locations_with_intersects)/RGC_premium_df["Address"].count(), len(num_locations_with_intersects)

def RGC_percent_in_rad_premium_Both(num_loc,rad, premium):

    num_locations_with_intersects = []
    for rgc_location in RGC_premium_df.iterrows():
        current_intersects = 0
        lst1 = AMC_premium_df.iterrows()
        lst2 = CNK_premium_df.iterrows()
        if not premium:
            lst1 = AMC_reg_df.iterrows()
            lst2 = CNK_reg_df.iterrows()
        for amc_location in lst1:
            if get_distance(rgc_location[1]["lats"],rgc_location[1]["longs"],amc_location[1]["lats"],amc_location[1]["longs"]) <= rad:
                for cnk_location in lst2:
                    if get_distance(cnk_location[1]["lats"],cnk_location[1]["longs"],rgc_location[1]["lats"],rgc_location[1]["longs"]) <= rad:
                        current_intersects += 1
            if current_intersects == num_loc:
                num_locations_with_intersects.append(rgc_location[1]["Address"])
                break
    return len(num_locations_with_intersects)/RGC_premium_df["Address"].count(), len(num_locations_with_intersects)

def RGC_percent_in_rad_premium_None(num_loc,rad, premium):

    num_locations_with_intersects = []
    for rgc_location in RGC_premium_df.iterrows():
        current_intersects = 0
        lst1 = AMC_premium_df.iterrows()
        lst2 = CNK_premium_df.iterrows()
        if not premium:
            lst1 = AMC_reg_df.iterrows()
            lst2 = CNK_reg_df.iterrows()
        for amc_location in lst1:
            if not get_distance(rgc_location[1]["lats"],rgc_location[1]["longs"],amc_location[1]["lats"],amc_location[1]["longs"]) <= rad:
                for cnk_location in lst2:
                    if not get_distance(cnk_location[1]["lats"],cnk_location[1]["longs"],rgc_location[1]["lats"],rgc_location[1]["longs"]) <= rad:
                        current_intersects += 1
            if current_intersects == num_loc:
                num_locations_with_intersects.append(rgc_location[1]["Address"])
                break
    return len(num_locations_with_intersects)/RGC_premium_df["Address"].count(), len(num_locations_with_intersects)


def CNK_percent_in_rad_premium(num_loc,rad):

    num_locations_with_intersects = []
    for cnk_location in CNK_premium_df.iterrows():
        current_intersects = 0
        check_rgc = True
        for amc_location in AMC_premium_df.iterrows():
            if get_distance(cnk_location[1]["lats"],cnk_location[1]["longs"],amc_location[1]["lats"],amc_location[1]["longs"]) <= rad:
                current_intersects += 1
            if current_intersects == num_loc:
                num_locations_with_intersects.append(cnk_location[1]["Address"])
                check_rgc = False
                break
        if check_rgc:
            for rgc_location in RGC_premium_df.iterrows():
                if get_distance(cnk_location[1]["lats"],cnk_location[1]["longs"],rgc_location[1]["lats"],rgc_location[1]["longs"]) <= rad:
                    current_intersects += 1
                if current_intersects == num_loc:
                    num_locations_with_intersects.append(cnk_location[1]["Address"])
                    break
    return len(num_locations_with_intersects)/CNK_premium_df["Address"].count(), len(num_locations_with_intersects)


#Functions calculate the percent of premium locations that have at least "num_loc"
#number of non premium location from any company in a "rad" mile radius
def AMC_percent_in_rad_reg(num_loc,rad):

    num_locations_with_intersects = []
    for amc_location in AMC_premium_df.iterrows():
        current_intersects = 0
        check_cnk = True
        for rgc_location in RGC_reg_df.iterrows():
            if get_distance(rgc_location[1]["lats"],rgc_location[1]["longs"],amc_location[1]["lats"],amc_location[1]["longs"]) <= rad:
                current_intersects += 1
            if current_intersects == num_loc:
                num_locations_with_intersects.append(amc_location[1]["Address"])
                check_cnk = False
                break
        if check_cnk:
            for cnk_location in CNK_reg_df.iterrows():
                if get_distance(cnk_location[1]["lats"],cnk_location[1]["longs"],amc_location[1]["lats"],amc_location[1]["longs"]) <= rad:
                    current_intersects += 1
                if current_intersects == num_loc:
                    num_locations_with_intersects.append(amc_location[1]["Address"])
                    break
    return len(num_locations_with_intersects)/AMC_premium_df["Address"].count(), len(num_locations_with_intersects)

def RGC_percent_in_rad_reg(num_loc,rad):

    num_locations_with_intersects = []
    for rgc_location in RGC_premium_df.iterrows():
        current_intersects = 0
        check_cnk = True
        for amc_location in AMC_reg_df.iterrows():
            if get_distance(rgc_location[1]["lats"],rgc_location[1]["longs"],amc_location[1]["lats"],amc_location[1]["longs"]) <= rad:
                current_intersects += 1
            if current_intersects == num_loc:
                num_locations_with_intersects.append(rgc_location[1]["Address"])
                check_cnk = False
                break
        if check_cnk:
            for cnk_location in CNK_reg_df.iterrows():
                if get_distance(cnk_location[1]["lats"],cnk_location[1]["longs"],rgc_location[1]["lats"],rgc_location[1]["longs"]) <= rad:
                    current_intersects += 1
                if current_intersects == num_loc:
                    num_locations_with_intersects.append(rgc_location[1]["Address"])
                    break
    return len(num_locations_with_intersects)/RGC_premium_df["Address"].count(), len(num_locations_with_intersects)

def CNK_percent_in_rad_premium_AMC(num_loc,rad, premium):

    num_locations_with_intersects = []
    for cnk_location in CNK_premium_df.iterrows():
        current_intersects = 0
        lst = AMC_premium_df.iterrows()
        if not premium:
            lst = AMC_reg_df.iterrows()
        for amc_location in lst:
            if get_distance(cnk_location[1]["lats"],cnk_location[1]["longs"],amc_location[1]["lats"],amc_location[1]["longs"]) <= rad:
                current_intersects += 1
            if current_intersects == num_loc:
                num_locations_with_intersects.append(cnk_location[1]["Address"])
                break
    return len(num_locations_with_intersects)/CNK_premium_df["Address"].count(), len(num_locations_with_intersects)

def CNK_percent_in_rad_premium_RGC(num_loc,rad, premium):

    num_locations_with_intersects = []
    for cnk_location in CNK_premium_df.iterrows():
        current_intersects = 0
        lst = RGC_premium_df.iterrows()
        if not premium:
            lst = RGC_reg_df.iterrows()
        for rgc_location in lst:
            if get_distance(cnk_location[1]["lats"],cnk_location[1]["longs"],rgc_location[1]["lats"],rgc_location[1]["longs"]) <= rad:
                current_intersects += 1
            if current_intersects == num_loc:
                num_locations_with_intersects.append(cnk_location[1]["Address"])
                break
    return len(num_locations_with_intersects)/CNK_premium_df["Address"].count(), len(num_locations_with_intersects)

def CNK_percent_in_rad_premium_Both(num_loc,rad, premium):

    num_locations_with_intersects = []
    for cnk_location in CNK_premium_df.iterrows():
        current_intersects = 0
        lst1 = AMC_premium_df.iterrows()
        lst2 = RGC_premium_df.iterrows()
        if not premium:
            lst1 = AMC_reg_df.iterrows()
            lst2 = RGC_reg_df.iterrows()
        for amc_location in lst1:
            if get_distance(cnk_location[1]["lats"],cnk_location[1]["longs"],amc_location[1]["lats"],amc_location[1]["longs"]) <= rad:
                for rgc_location in lst2:
                    if get_distance(cnk_location[1]["lats"],cnk_location[1]["longs"],rgc_location[1]["lats"],rgc_location[1]["longs"]) <= rad:
                        current_intersects += 1
            if current_intersects == num_loc:
                num_locations_with_intersects.append(cnk_location[1]["Address"])
                break
    return len(num_locations_with_intersects)/CNK_premium_df["Address"].count(), len(num_locations_with_intersects)

def CNK_percent_in_rad_premium_None(num_loc,rad, premium):

    num_locations_with_intersects = []
    for cnk_location in CNK_premium_df.iterrows():
        current_intersects = 0
        lst1 = AMC_premium_df.iterrows()
        lst2 = RGC_premium_df.iterrows()
        if not premium:
            lst1 = AMC_reg_df.iterrows()
            lst2 = RGC_reg_df.iterrows()
        for amc_location in lst1:
            if not get_distance(cnk_location[1]["lats"],cnk_location[1]["longs"],amc_location[1]["lats"],amc_location[1]["longs"]) <= rad:
                for rgc_location in lst2:
                    if not get_distance(cnk_location[1]["lats"],cnk_location[1]["longs"],rgc_location[1]["lats"],rgc_location[1]["longs"]) <= rad:
                        current_intersects += 1
            if current_intersects == num_loc:
                num_locations_with_intersects.append(cnk_location[1]["Address"])
                break
    return len(num_locations_with_intersects)/CNK_premium_df["Address"].count(), len(num_locations_with_intersects)


def CNK_percent_in_rad_reg(num_loc,rad):

    num_locations_with_intersects = []
    for cnk_location in CNK_premium_df.iterrows():
        current_intersects = 0
        check_rgc = True
        for amc_location in AMC_reg_df.iterrows():
            if get_distance(cnk_location[1]["lats"],cnk_location[1]["longs"],amc_location[1]["lats"],amc_location[1]["longs"]) <= rad:
                current_intersects += 1
            if current_intersects == num_loc:
                num_locations_with_intersects.append(cnk_location[1]["Address"])
                check_rgc = False
                break
        if check_rgc:
            for rgc_location in RGC_reg_df.iterrows():
                if get_distance(cnk_location[1]["lats"],cnk_location[1]["longs"],rgc_location[1]["lats"],rgc_location[1]["longs"]) <= rad:
                    current_intersects += 1
                if current_intersects == num_loc:
                    num_locations_with_intersects.append(cnk_location[1]["Address"])
                    break

    return len(num_locations_with_intersects)/CNK_premium_df["Address"].count(), len(num_locations_with_intersects)

#Functions calculate the percent of premium locations that have at least "num_loc"
#number of all locations from any company in a "rad" mile radius
def AMC_percent_in_rad_all(num_loc,rad):

    num_locations_with_intersects = []
    for amc_location in AMC_premium_df.iterrows():
        current_intersects = 0
        check_cnk = True
        for rgc_location in RGC_all_df.iterrows():
            if get_distance(rgc_location[1]["lats"],rgc_location[1]["longs"],amc_location[1]["lats"],amc_location[1]["longs"]) <= rad:
                current_intersects += 1
            if current_intersects == num_loc:
                num_locations_with_intersects.append(amc_location[1]["Address"])
                check_cnk = False
                break
        if check_cnk:
            for cnk_location in CNK_all_df.iterrows():
                if get_distance(cnk_location[1]["lats"],cnk_location[1]["longs"],amc_location[1]["lats"],amc_location[1]["longs"]) <= rad:
                    current_intersects += 1
                if current_intersects == num_loc:
                    num_locations_with_intersects.append(amc_location[1]["Address"])
                    break
    return len(num_locations_with_intersects)/AMC_premium_df["Address"].count(), len(num_locations_with_intersects)

def RGC_percent_in_rad_all(num_loc,rad):

    num_locations_with_intersects = []
    for rgc_location in RGC_premium_df.iterrows():
        current_intersects = 0
        check_cnk = True
        for amc_location in AMC_all_df.iterrows():
            if get_distance(rgc_location[1]["lats"],rgc_location[1]["longs"],amc_location[1]["lats"],amc_location[1]["longs"]) <= rad:
                current_intersects += 1
            if current_intersects == num_loc:
                num_locations_with_intersects.append(rgc_location[1]["Address"])
                check_cnk = False
                break
        if check_cnk:
            for cnk_location in CNK_all_df.iterrows():
                if get_distance(cnk_location[1]["lats"],cnk_location[1]["longs"],rgc_location[1]["lats"],rgc_location[1]["longs"]) <= rad:
                    current_intersects += 1
                if current_intersects == num_loc:
                    num_locations_with_intersects.append(rgc_location[1]["Address"])
                    break
    return len(num_locations_with_intersects)/RGC_premium_df["Address"].count(), len(num_locations_with_intersects)

def CNK_percent_in_rad_all(num_loc,rad):

    num_locations_with_intersects = []
    for cnk_location in CNK_premium_df.iterrows():
        current_intersects = 0
        check_rgc = True
        for amc_location in AMC_all_df.iterrows():
            if get_distance(cnk_location[1]["lats"],cnk_location[1]["longs"],amc_location[1]["lats"],amc_location[1]["longs"]) <= rad:
                current_intersects += 1
            if current_intersects == num_loc:
                num_locations_with_intersects.append(cnk_location[1]["Address"])
                check_rgc = False
                break
        if check_rgc:
            for rgc_location in RGC_all_df.iterrows():
                if get_distance(cnk_location[1]["lats"],cnk_location[1]["longs"],rgc_location[1]["lats"],rgc_location[1]["longs"]) <= rad:
                    current_intersects += 1
                if current_intersects == num_loc:
                    num_locations_with_intersects.append(cnk_location[1]["Address"])
                    break

    return len(num_locations_with_intersects)/CNK_premium_df["Address"].count(), len(num_locations_with_intersects)




#Analysis output to FILENAME, modify MIN_LOCATIONS and RAD as needed

MIN_LOCATIONS = 1 #minimum muber of locations for the theatre to be considered intersected
RAD = 15 #in miles
FILENAME = "Theater Analysis Output Template.xlsx"




wb = load_workbook(FILENAME)
ws = wb['Sheet1']
ws['B7'] = AMC_percent_in_rad_premium_RGC(MIN_LOCATIONS,RAD, True)[0]
ws['C7'] = AMC_percent_in_rad_premium_CNK(MIN_LOCATIONS,RAD, True)[0]
ws['D7'] = AMC_percent_in_rad_premium_Both(MIN_LOCATIONS,RAD, True)[0]
ws['E7'] = 1-AMC_percent_in_rad_premium(MIN_LOCATIONS,RAD)[0]

ws['B8'] = AMC_percent_in_rad_premium_RGC(MIN_LOCATIONS,RAD, True)[1]
ws['C8'] = AMC_percent_in_rad_premium_CNK(MIN_LOCATIONS,RAD, True )[1]
ws['D8'] = AMC_percent_in_rad_premium_Both(MIN_LOCATIONS,RAD, True)[1]
ws['E8'] = 1-AMC_percent_in_rad_premium(MIN_LOCATIONS,RAD)[1]

ws['G7'] = AMC_percent_in_rad_premium_RGC(MIN_LOCATIONS,RAD, False)[0]
ws['H7'] = AMC_percent_in_rad_premium_CNK(MIN_LOCATIONS,RAD, False)[0]
ws['I7'] = AMC_percent_in_rad_premium_Both(MIN_LOCATIONS,RAD, False)[0]
ws['J7'] = 1-AMC_percent_in_rad_reg(MIN_LOCATIONS,RAD)[0]

ws['G8'] = AMC_percent_in_rad_premium_RGC(MIN_LOCATIONS,RAD, False)[1]
ws['H8'] = AMC_percent_in_rad_premium_CNK(MIN_LOCATIONS,RAD, False )[1]
ws['I8'] = AMC_percent_in_rad_premium_Both(MIN_LOCATIONS,RAD, False)[1]
ws['J8'] = 1-AMC_percent_in_rad_reg(MIN_LOCATIONS,RAD)[1]

#For RGC
ws['B14'] = RGC_percent_in_rad_premium_AMC(MIN_LOCATIONS,RAD, True)[0]
ws['C14'] = RGC_percent_in_rad_premium_CNK(MIN_LOCATIONS,RAD, True)[0]
ws['D14'] = RGC_percent_in_rad_premium_Both(MIN_LOCATIONS,RAD, True)[0]
ws['E14'] = 1-RGC_percent_in_rad_premium(MIN_LOCATIONS,RAD)[0]

ws['B15'] = RGC_percent_in_rad_premium_AMC(MIN_LOCATIONS,RAD, True)[1]
ws['C15'] = RGC_percent_in_rad_premium_CNK(MIN_LOCATIONS,RAD, True )[1]
ws['D15'] = RGC_percent_in_rad_premium_Both(MIN_LOCATIONS,RAD, True)[1]
ws['E15'] = 1-RGC_percent_in_rad_premium(MIN_LOCATIONS,RAD)[1]

ws['G14'] = RGC_percent_in_rad_premium_AMC(MIN_LOCATIONS,RAD, False)[0]
ws['H14'] = RGC_percent_in_rad_premium_CNK(MIN_LOCATIONS,RAD, False)[0]
ws['I14'] = RGC_percent_in_rad_premium_Both(MIN_LOCATIONS,RAD, False)[0]
ws['J14'] = 1-RGC_percent_in_rad_reg(MIN_LOCATIONS,RAD)[0]

ws['G15'] = RGC_percent_in_rad_premium_AMC(MIN_LOCATIONS,RAD, False)[1]
ws['H15'] = RGC_percent_in_rad_premium_CNK(MIN_LOCATIONS,RAD, False )[1]
ws['I15'] = RGC_percent_in_rad_premium_Both(MIN_LOCATIONS,RAD, False)[1]
ws['J15'] = 1-RGC_percent_in_rad_reg(MIN_LOCATIONS,RAD)[1]

#For CNK
ws['B21'] = CNK_percent_in_rad_premium_RGC(MIN_LOCATIONS,RAD, True)[0]
ws['C21'] = CNK_percent_in_rad_premium_AMC(MIN_LOCATIONS,RAD, True)[0]
ws['D21'] = CNK_percent_in_rad_premium_Both(MIN_LOCATIONS,RAD, True)[0]
ws['E21'] =  1-CNK_percent_in_rad_premium(MIN_LOCATIONS,RAD)[0]

ws['B22'] = CNK_percent_in_rad_premium_RGC(MIN_LOCATIONS,RAD, True)[1]
ws['C22'] = CNK_percent_in_rad_premium_AMC(MIN_LOCATIONS,RAD, True )[1]
ws['D22'] = CNK_percent_in_rad_premium_Both(MIN_LOCATIONS,RAD, True)[1]
ws['E22'] =  1-CNK_percent_in_rad_premium(MIN_LOCATIONS,RAD)[1]

ws['G21'] = CNK_percent_in_rad_premium_RGC(MIN_LOCATIONS,RAD, False)[0]
ws['H21'] = CNK_percent_in_rad_premium_AMC(MIN_LOCATIONS,RAD, False)[0]
ws['I21'] = CNK_percent_in_rad_premium_Both(MIN_LOCATIONS,RAD, False)[0]
ws['J21'] = 1-CNK_percent_in_rad_reg(MIN_LOCATIONS,RAD)[0]

ws['G22'] = CNK_percent_in_rad_premium_RGC(MIN_LOCATIONS,RAD, False)[1]
ws['H22'] = CNK_percent_in_rad_premium_AMC(MIN_LOCATIONS,RAD, False )[1]
ws['I22'] = CNK_percent_in_rad_premium_Both(MIN_LOCATIONS,RAD, False)[1]
ws['J22'] = 1-CNK_percent_in_rad_reg(MIN_LOCATIONS,RAD)[1]


wb.save(FILENAME)
#ws.cell("C8").value = 3
# base_df = pd.DataFrame(columns=['Tickers','Premium Intersects', 'Regular Intersects', 'All Intersects'])
# base_df['Tickers'] = ["AMC", "RGC", "CNK"]
# base_df['Premium Intersects'] = [AMC_percent_in_rad_premium(MIN_LOCATIONS,RAD),RGC_percent_in_rad_premium(MIN_LOCATIONS,RAD),CNK_percent_in_rad_premium(MIN_LOCATIONS,RAD)]
# base_df['Regular Intersects'] = [AMC_percent_in_rad_reg(MIN_LOCATIONS,RAD),RGC_percent_in_rad_reg(MIN_LOCATIONS,RAD),CNK_percent_in_rad_reg(MIN_LOCATIONS,RAD)]
# base_df['All Intersects'] = [AMC_percent_in_rad_all(MIN_LOCATIONS,RAD),RGC_percent_in_rad_all(MIN_LOCATIONS,RAD),CNK_percent_in_rad_all(MIN_LOCATIONS,RAD)]
#
#
# writer = pd.ExcelWriter(FILENAME)
# base_df.to_excel(writer,'Data')
# writer.save()
